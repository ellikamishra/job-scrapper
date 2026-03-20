#!/usr/bin/env python3
"""Create a sample input Excel file for the job scraper."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side

wb = Workbook()
ws = wb.active
ws.title = "Companies"

header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                     top=Side(style="thin"), bottom=Side(style="thin"))

for col, name in enumerate(["Company Name", "Domain", "Career URL"], 1):
    cell = ws.cell(row=1, column=col, value=name)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border

companies = [
    ("Google", "google.com", "https://www.google.com/about/careers/applications/jobs/results"),
    ("Microsoft", "microsoft.com", "https://careers.microsoft.com/us/en/search-results"),
    ("Amazon", "amazon.com", "https://www.amazon.jobs/en/search"),
    ("Netflix", "netflix.com", "https://jobs.netflix.com/search"),
    ("Stripe", "stripe.com", "https://stripe.com/jobs/search"),
    ("Uber", "uber.com", "https://www.uber.com/us/en/careers/list/"),
    ("Meta", "meta.com", "https://www.metacareers.com/jobs"),
    ("Apple", "apple.com", "https://jobs.apple.com/en-us/search"),
    ("Spotify", "spotify.com", "https://www.lifeatspotify.com/jobs"),
    ("Atlassian", "atlassian.com", "https://www.atlassian.com/company/careers/all-jobs"),
]

for row, (name, domain, career_url) in enumerate(companies, 2):
    ws.cell(row=row, column=1, value=name).border = thin_border
    ws.cell(row=row, column=2, value=domain).border = thin_border
    ws.cell(row=row, column=3, value=career_url).border = thin_border

ws.column_dimensions["A"].width = 20
ws.column_dimensions["B"].width = 25
ws.column_dimensions["C"].width = 60
wb.save("sample_companies.xlsx")
print("[+] Created sample_companies.xlsx")
