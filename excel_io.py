"""
Excel reading and writing utilities.
"""
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def read_companies(filepath: str) -> list[dict]:
    if not os.path.exists(filepath):
        raise FileNotFoundError(f"Input file not found: {filepath}")

    wb = load_workbook(filepath, read_only=True)
    ws = wb.active

    headers = []
    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        val = str(cell.value).strip().lower() if cell.value else ""
        headers.append(val)

    companies = []
    for row in ws.iter_rows(min_row=2):
        values = [cell.value for cell in row]
        if not any(values):
            continue

        entry = {}
        for i, header in enumerate(headers):
            if i < len(values):
                entry[header] = str(values[i]).strip() if values[i] else ""

        name = (entry.get("company name") or entry.get("company") or
                entry.get("name") or "")
        domain = (entry.get("domain") or entry.get("website") or
                  entry.get("url") or "")
        career_url = (entry.get("career url") or entry.get("career_url") or
                      entry.get("careers url") or entry.get("career page") or "")

        if not name:
            continue

        if not domain:
            slug = name.lower().replace(" ", "").replace(".", "").replace(",", "")
            domain = f"{slug}.com"

        domain = domain.replace("https://", "").replace("http://", "")
        domain = domain.replace("www.", "").rstrip("/")

        companies.append({
            "name": name,
            "domain": domain,
            "career_url": career_url,
        })

    wb.close()
    return companies


def write_results(filepath: str, jobs: list[dict]):
    wb = Workbook()
    ws = wb.active
    ws.title = "Job Listings"

    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496",
                              fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    link_font = Font(color="0563C1", underline="single")

    columns = ["Company Name", "Job Title", "Location", "Experience",
               "Skills Matched", "Link"]
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for row_idx, job in enumerate(jobs, 2):
        ws.cell(row=row_idx, column=1, value=job.get("company", "")).border = thin_border
        ws.cell(row=row_idx, column=2, value=job.get("title", "")).border = thin_border
        ws.cell(row=row_idx, column=3, value=job.get("location", "N/A")).border = thin_border
        ws.cell(row=row_idx, column=4, value=job.get("experience", "Not specified")).border = thin_border
        ws.cell(row=row_idx, column=5, value=job.get("skills_matched", "")).border = thin_border

        link_cell = ws.cell(row=row_idx, column=6, value=job.get("link", ""))
        link_cell.font = link_font
        link_cell.border = thin_border
        if job.get("link"):
            link_cell.hyperlink = job["link"]

    col_widths = [25, 50, 25, 20, 30, 60]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width

    ws.freeze_panes = "A2"
    wb.save(filepath)
    print(f"\n[+] Results saved to: {filepath}")
    print(f"[+] Total jobs found: {len(jobs)}")
